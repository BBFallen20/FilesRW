import csv
import openpyxl
import xlsxwriter


class ReadHandler:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_extension = self.file_path.split('.')[-1] if self.file_path.__contains__('.') else None

    def read_txt(self) -> bool:
        with open(self.file_path, 'r') as File:
            if len(File.readline(1)) > 0:
                print('Here is first lines from the file:\n')
                [print(line) for line in File.readlines()[:5]]
                return True
            print('File is empty.')
            return False

    def read_csv(self) -> bool:
        with open(self.file_path) as csv_file:
            reader = csv.reader(csv_file, delimiter=' ',
                                quotechar='|', quoting=csv.QUOTE_MINIMAL)
            counter = 0
            for row in reader:
                print(row)
                counter += 1
                if counter == 5:
                    break
        return True

    def read_xlsx(self) -> bool:
        wb_obj = openpyxl.load_workbook(self.file_path)
        sheet = wb_obj.active
        for row in sheet.iter_rows(max_row=5):
            for cell in row:
                print(cell.value, end=" ")
            print()
        return True

    def read(self) -> bool:
        """Main read method which calls another reading methods depending on the file extension"""
        extensions = {
            'txt': self.read_txt,
            'csv': self.read_csv,
            'xlsx': self.read_xlsx,
            'xls': self.read_txt
        }
        if self.file_extension in extensions:
            try:
                return extensions[self.file_extension]()
            except FileNotFoundError:
                print('File not found.')
            except PermissionError:
                print('You don`t have permission to use this file.')
            except IsADirectoryError:
                print('Path is a directory.')
            return False


class WriteHandler:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_extension = file_path.split('.')[-1]

    def write_txt_file(self) -> bool:
        with open(self.file_path, 'a') as File:
            File.writelines([input('Enter line:\n') + '\n' for _ in range(int(input('Enter lines amount:\n')))])
        print('File extended successfully.')
        return True

    def write_csv_file(self) -> bool:
        with open(self.file_path, 'a', newline='') as csv_file:
            writer = csv.writer(csv_file, delimiter=' ',
                                quotechar='|', quoting=csv.QUOTE_MINIMAL)
            writer.writerow([input('Enter line:\n') + ';' for _ in range(int(input('Enter lines amount:\n')))])
        print('File extended successfully.')
        return True

    def write_xlsx_file(self) -> bool:
        workbook = xlsxwriter.Workbook(self.file_path)
        worksheet = workbook.add_worksheet('test')
        data_list = [input('Enter column:\n') for _ in range(int(input('Enter columns amount:\n')))]
        for col_num, data in enumerate(data_list):
            worksheet.write(0, col_num, data)
        workbook.close()
        print('File extended successfully.')
        return True

    def write(self) -> bool:
        """Main write method which calls another writing methods depending on the file extension"""
        extensions = {
            'txt': self.write_txt_file,
            'csv': self.write_csv_file,
            'xlsx': self.write_xlsx_file,
            'xls': self.write_xlsx_file
        }
        if self.file_extension in extensions:
            try:
                return extensions[self.file_extension]()
            except FileNotFoundError:
                print('File not found.')
            except PermissionError:
                print('You don`t have permission to use this file.')
            except ValueError:
                print('Enter digit for lines amount.')
            except IsADirectoryError:
                print('Path is a directory.')
            return False
        return False
