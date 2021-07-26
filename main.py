import csv
import argparse
import openpyxl as openpyxl
import xlsxwriter


def read_txt(file_path: str) -> bool:
    with open(file_path, 'r') as File:
        if len(File.readline(1)) > 0:
            print('Here is first lines from the file:\n')
            [print(line) for line in File.readlines()[:5]]
            return True
        print('File is empty.')
        return False


def read_csv(file_path: str) -> bool:
    with open(file_path) as csv_file:
        reader = csv.reader(csv_file, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
        counter = 0
        for row in reader:
            print(row)
            counter += 1
            if counter == 5:
                break
    return True


def read_xlsx(file_path: str) -> bool:
    wb_obj = openpyxl.load_workbook(file_path)

    # Read the active sheet:
    sheet = wb_obj.active
    for row in sheet.iter_rows(max_row=5):
        for cell in row:
            print(cell.value, end=" ")
        print()
    return True


def read_mode(file_path: str) -> bool:
    extensions = {'txt': read_txt, 'csv': read_csv, 'xlsx': read_xlsx, 'xls': read_txt}
    path, extension = file_path.split('.') if file_path.__contains__('.') else [None, None]
    if extension in extensions:
        try:
            return extensions[extension](file_path)
        except FileNotFoundError:
            print('File not found.')
        except PermissionError:
            print('You don`t have permission to use this file.')
        except IsADirectoryError:
            print('Path is a directory.')
        return False


def write_txt_file(file_path: str) -> bool:
    with open(file_path, 'a') as File:
        File.writelines([input('Enter line:\n') + '\n' for _ in range(int(input('Enter lines amount:\n')))])
    print('File extended successfully.')
    return True


def write_csv_file(file_path: str) -> bool:
    with open(file_path, 'a', newline='') as csv_file:
        writer = csv.writer(csv_file, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
        writer.writerow([input('Enter line:\n') + ';' for _ in range(int(input('Enter lines amount:\n')))])
    print('File extended successfully.')
    return True


def write_xlsx_file(file_path: str) -> bool:
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet('test')
    data_list = [input('Enter column:\n') for _ in range(int(input('Enter columns amount:\n')))]
    for col_num, data in enumerate(data_list):
        worksheet.write(0, col_num, data)
    workbook.close()
    print('File extended successfully.')
    return True


def write_mode(file_path: str) -> bool:
    extensions = {'txt': write_txt_file, 'csv': write_csv_file, 'xlsx': write_xlsx_file, 'xls': write_xlsx_file}
    path, extension = file_path.split('.') if file_path.__contains__('.') else [None, None]
    if extension in extensions:
        try:
            return extensions[extension](file_path)
        except FileNotFoundError:
            print('File not found.')
        except PermissionError:
            print('You don`t have permission to use this file.')
        except ValueError:
            print('Enter digit for lines amount.')
        except IsADirectoryError:
            print('Path is a directory.')
        return False
    return False


def choose_mode(mode: str) -> bool:
    modes = {'read': read_mode, 'write': write_mode}
    completed = False
    if mode in modes:
        completed = modes[mode](input('Enter file path:\n'))
    else:
        print('Error. Mode does not exist.')
    return completed


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Application mode(write/read) selector')
    parser.add_argument(
        'mode',
        metavar='mode',
        type=str,
        help='application mode'
    )
    args = parser.parse_args()
    selected_mode = args.mode.lower()
    choose_mode(selected_mode)
